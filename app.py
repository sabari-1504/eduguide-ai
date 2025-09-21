import tkinter as tk
from tkinter import messagebox
import os
from openpyxl import Workbook, load_workbook
from flask_cors import CORS

EXCEL_FILE = 'users.xlsx'

# Ensure Excel file exists with headers
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['username', 'password'])
    wb.save(EXCEL_FILE)

def register_user(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    # Check if username already exists
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return False, 'Username already exists.'
    ws.append([username, password])
    wb.save(EXCEL_FILE)
    return True, 'Registration successful!'

def login_user(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return True, 'Login successful!'
    return False, 'Invalid username or password.'

class AuthApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Login/Register App')
        self.root.geometry('300x200')
        self.show_login()

    def clear_frame(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def show_login(self):
        self.clear_frame()
        tk.Label(self.root, text='Login', font=('Arial', 16)).pack(pady=10)
        tk.Label(self.root, text='Username').pack()
        username_entry = tk.Entry(self.root)
        username_entry.pack()
        tk.Label(self.root, text='Password').pack()
        password_entry = tk.Entry(self.root, show='*')
        password_entry.pack()
        def attempt_login():
            username = username_entry.get().strip()
            password = password_entry.get().strip()
            success, msg = login_user(username, password)
            if success:
                messagebox.showinfo('Success', msg)
            else:
                messagebox.showerror('Error', msg)
        tk.Button(self.root, text='Login', command=attempt_login).pack(pady=5)
        tk.Button(self.root, text='Go to Register', command=self.show_register).pack()

    def show_register(self):
        self.clear_frame()
        tk.Label(self.root, text='Register', font=('Arial', 16)).pack(pady=10)
        tk.Label(self.root, text='Username').pack()
        username_entry = tk.Entry(self.root)
        username_entry.pack()
        tk.Label(self.root, text='Password').pack()
        password_entry = tk.Entry(self.root, show='*')
        password_entry.pack()
        def attempt_register():
            username = username_entry.get().strip()
            password = password_entry.get().strip()
            if not username or not password:
                messagebox.showerror('Error', 'Please fill all fields.')
                return
            success, msg = register_user(username, password)
            if success:
                messagebox.showinfo('Success', msg)
                self.show_login()
            else:
                messagebox.showerror('Error', msg)
        tk.Button(self.root, text='Register', command=attempt_register).pack(pady=5)
        tk.Button(self.root, text='Go to Login', command=self.show_login).pack()

if __name__ == '__main__':
    root = tk.Tk()
    app = AuthApp(root)
    root.mainloop() 